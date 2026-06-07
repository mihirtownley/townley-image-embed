import gc
import base64
import io
import logging
import threading
import requests
import os
from datetime import datetime, date
from flask import Flask, request, jsonify
from openpyxl import load_workbook
from PIL import Image as PILImage
import xlsxwriter

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)

STYLE_COL_IDX   = 0
IMAGE_COL_WIDTH = 18          # ✅ column width (chars)
ROW_HEIGHT_PT   = 75          # ✅ row height (points)
IMAGE_URL_BASE  = "https://app.townleygirl.com/Image/preview/"
REQUEST_TIMEOUT = 6

# ── Image quality settings ────────────────────────────────────────────────────
# We download the image directly from the source URL and embed the raw bytes.
# No re-encoding to PNG → file size stays small (each image ~3-5 KB JPEG).
# XlsxWriter's object_position=1 then scales it to fit inside the cell at
# render time, so it always looks sharp regardless of zoom level.
# ─────────────────────────────────────────────────────────────────────────────

def download_full_quality(style_str):
    """Download image from URL and return the raw JPEG bytes (no re-encoding)."""
    try:
        resp = requests.get(f"{IMAGE_URL_BASE}{style_str}", timeout=REQUEST_TIMEOUT, stream=True)
        resp.raise_for_status()

        data = resp.content
        del resp

        # Detect format and log dimensions without re-encoding
        with PILImage.open(io.BytesIO(data)) as img:
            fmt = img.format or "JPEG"
            w, h = img.size

        logging.info(f"IMG {style_str}: {len(data)/1024:.1f} KB, {fmt} {w}x{h}")
        return data

    except Exception as e:
        logging.warning(f"Image failed for {style_str}: {e}")
        return None
    finally:
        gc.collect()


def process_and_callback(excel_bytes, filename, callback_url):
    try:
        logging.info(f"Received file size: {len(excel_bytes)} bytes")

        # ── Read source data with openpyxl ────────────────────────────────────
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        del excel_bytes
        gc.collect()

        if ws['A1'].value and str(ws['A1'].value).strip().lower() == "image":
            logging.warning("File already processed — skipping")
            return

        headers  = [cell.value for cell in ws[1]]
        num_cols = len(headers)

        all_rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if any(v is not None for v in row):
                all_rows.append(list(row))

        del wb
        gc.collect()

        # Sort A→Z by Style
        all_rows.sort(key=lambda r: (
            r[STYLE_COL_IDX] is None,
            str(r[STYLE_COL_IDX]).lower() if r[STYLE_COL_IDX] else ""
        ))
        logging.info(f"Read {len(all_rows)} rows, sorted A→Z ✅")

        # ── Download images ────────────────────────────────────────────────────
        logging.info("Downloading images (raw JPEG, no re-encoding)...")
        images = {}
        for row_data in all_rows:
            style_val = row_data[STYLE_COL_IDX]
            if style_val:
                style_str = str(style_val).strip()
                if style_str and style_str not in images:
                    images[style_str] = download_full_quality(style_str)

        processed = sum(1 for v in images.values() if v is not None)
        skipped   = sum(1 for v in images.values() if v is None)
        logging.info(f"Images: {processed} downloaded, {skipped} failed ✅")

        # ── Calculate column widths ────────────────────────────────────────────
        col_widths = []
        for col_idx in range(num_cols):
            max_len = len(str(headers[col_idx])) if headers[col_idx] else 0
            for row_data in all_rows:
                if col_idx < len(row_data) and row_data[col_idx] is not None:
                    max_len = max(max_len, len(str(row_data[col_idx])))
            col_widths.append(min(max_len + 4, 50))

        # ── Build xlsx with XlsxWriter ─────────────────────────────────────────
        logging.info("Building xlsx with XlsxWriter insert_image()...")

        tmp_path = "/tmp/xlsxwriter_output.xlsx"
        workbook = xlsxwriter.Workbook(tmp_path)
        sheet    = workbook.add_worksheet()

        bold_fmt = workbook.add_format({"bold": True,  "font_name": "Calibri", "font_size": 11})
        cell_fmt = workbook.add_format({"font_name": "Calibri", "font_size": 11})
        date_fmt = workbook.add_format({"font_name": "Calibri", "font_size": 11, "num_format": "mm/dd/yyyy"})
        num_fmt  = workbook.add_format({"font_name": "Calibri", "font_size": 11})

        # Header row
        sheet.write(0, 0, "Image", bold_fmt)
        for col_idx, header in enumerate(headers):
            sheet.write(0, col_idx + 1, header if header is not None else "", bold_fmt)

        # Column widths
        sheet.set_column(0, 0, IMAGE_COL_WIDTH)
        for col_idx, width in enumerate(col_widths):
            sheet.set_column(col_idx + 1, col_idx + 1, width)

        # ── Centering offsets ─────────────────────────────────────────────────
        # XlsxWriter column-width unit ≈ 7.5 px per char; row height 1 pt ≈ 1.333 px
        COL_WIDTH_PX  = IMAGE_COL_WIDTH * 7.5   # ≈ 135 px
        ROW_HEIGHT_PX = ROW_HEIGHT_PT  * 1.333  # ≈ 100 px

        # Target display size: fill the cell with a small padding margin
        DISPLAY_W = int(COL_WIDTH_PX)  - 4
        DISPLAY_H = int(ROW_HEIGHT_PX) - 4

        # Centre offsets so the image sits in the middle of the cell
        x_offset = max(0, int((COL_WIDTH_PX  - DISPLAY_W) / 2))
        y_offset = max(0, int((ROW_HEIGHT_PX - DISPLAY_H) / 2))
        # ─────────────────────────────────────────────────────────────────────

             # Data rows with images
        for row_idx, row_data in enumerate(all_rows, start=1):
            sheet.set_row(row_idx, ROW_HEIGHT_PT)

            style_val = row_data[STYLE_COL_IDX]

            if style_val:
                style_str = str(style_val).strip()
                img_bytes = images.get(style_str)

                if img_bytes:
                    try:
                        with PILImage.open(io.BytesIO(img_bytes)) as img:
                            img_w, img_h = img.size

                        scale = min(
                            DISPLAY_W / img_w,
                            DISPLAY_H / img_h
                        )

                        final_w = img_w * scale
                        final_h = img_h * scale

                        center_x = max(0, int((DISPLAY_W - final_w) / 2))
                        center_y = max(0, int((DISPLAY_H - final_h) / 2))

                        sheet.insert_image(
                            row_idx,
                            0,
                            "image.jpg",
                            {
                                "image_data": io.BytesIO(img_bytes),
                                "x_offset": center_x,
                                "y_offset": center_y,
                                "x_scale": scale,
                                "y_scale": scale,
                                "object_position": 1,
                            }
                        )

                    except Exception as e:
                        logging.warning(
                            f"insert_image failed for {style_str}: {e}"
                        )

            # Write row data
            for col_idx, value in enumerate(row_data):
                dest_col = col_idx + 1

                if value is None:
                    sheet.write_blank(row_idx, dest_col, None, cell_fmt)

                elif isinstance(value, bool):
                    sheet.write_boolean(row_idx, dest_col, value, cell_fmt)

                elif isinstance(value, (datetime, date)):
                    sheet.write_datetime(row_idx, dest_col, value, date_fmt)

                elif isinstance(value, (int, float)):
                    sheet.write_number(row_idx, dest_col, value, num_fmt)

                else:
                    sheet.write_string(row_idx, dest_col, str(value), cell_fmt)

                sheet.freeze_panes(1, 0)
        sheet.autofilter(0, 0, len(all_rows), num_cols)
        workbook.close()

        # Read from temp file
        with open(tmp_path, "rb") as f:
            output_bytes = f.read()

        try:
            os.remove(tmp_path)
        except Exception:
            pass

        if output_bytes.startswith(b'PK\x03\x04'):
            logging.info(f"Output xlsx VALID ✅ size: {len(output_bytes)/1024/1024:.2f} MB")
        else:
            logging.error("Output xlsx INVALID ❌")

        result_b64 = base64.b64encode(output_bytes).decode("utf-8")
        logging.info(f"Base64 length: {len(result_b64)}")
        del output_bytes
        gc.collect()

        if not filename.lower().endswith(".xlsx"):
            filename = filename + ".xlsx"

        payload = {
            "file"      : result_b64,
            "filename"  : filename,
            "processed" : processed,
            "skipped"   : skipped,
        }
        resp = requests.post(callback_url, json=payload, timeout=30)
        logging.info(f"Callback: {resp.status_code}, processed={processed}, skipped={skipped}")

    except Exception as e:
        logging.error(f"Background processing failed: {e}", exc_info=True)


@app.route("/embed_images", methods=["POST"])
def embed_images():
    try:
        body         = request.get_json(force=True)
        excel_b64    = body.get("file")
        filename     = body.get("filename", "report.xlsx")
        callback_url = os.environ.get("CALLBACK_URL")
    except Exception as e:
        return jsonify({"error": f"Invalid request: {e}"}), 400

    if not excel_b64:
        return jsonify({"error": "Missing 'file' field"}), 400
    if not callback_url:
        return jsonify({"error": "CALLBACK_URL not configured"}), 500

    try:
        excel_bytes = base64.b64decode(excel_b64)
        logging.info(f"Decoded excel_bytes size: {len(excel_bytes)} bytes")

        for attempt in range(4):
            if excel_bytes.startswith(b'PK\x03\x04'):
                logging.info(f"Valid xlsx after {attempt + 1} decode(s), size: {len(excel_bytes)} bytes")
                break
            logging.info(f"Attempt {attempt + 1}: starts with {excel_bytes[:4]}, decoding again...")
            excel_bytes = base64.b64decode(excel_bytes)
        else:
            return jsonify({"error": "Could not decode xlsx after 4 attempts"}), 422

    except Exception as e:
        return jsonify({"error": f"Base64 decode failed: {e}"}), 422

    thread = threading.Thread(
        target=process_and_callback,
        args=(excel_bytes, filename, callback_url),
        daemon=True
    )
    thread.start()

    return jsonify({"status": "accepted", "message": "Processing started"}), 202


@app.route("/", methods=["GET"])
def health_check():
    return jsonify({
        "status"  : "Townley Image Embedder is Live",
        "endpoint": "/embed_images (POST only)"
    }), 200


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
